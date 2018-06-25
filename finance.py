from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import math
import urllib.request,urllib.parse, urllib.error
import ssl
from ClassStock import Stock
import json
import os, errno
import os.path, time
import sys
import datetime
import numpy
from pytz import timezone
from datetime import date, timedelta
from shutil import copyfile
import smtplib

#TODO TODO 1. Automate daily run and inform me when valuation changes happen
#TODO refactor code break up the main into different pieces
#TODO nice to have: put focus back on the result sheet instead of port sheet
#TODO nice to have: compare current snapshot with historical data, might be able to try this with tableau

COLUMN_VALUATION = 'Z'
COLUMN_VALUATION_NUMBER = 26
COLUMN_NAME = 'A'
COLUMN_SYMBOL = 'B'
COLUMN_CUR_PRICE = 'C'
COLUMN_FAIR_PRICE = 'D'
COLUMN_BUSINESS_VALUE = 'E'
COLUMN_CASH_VALUE = 'F'
COLUMN_THESIS = 'G'
COLUMN_UPSIDE = 'H'
COLUMN_CUSHION = 'I'
SHEET_NAME = "Sheet1"
PORT_SHEET_NAME = "Port"
RESULT_SHEET_NAME = "Result"
PORT_COLUMN_SYMBOL = 'A'
BATCH_SIZE = 100
UPSIDE_ALERT = 0.9
DOWNSIDE_ALERT = 0.15
#count number of rows
symbolsCount = -1#offset the header count
numOfPos  = -1
PORT_CELL_UNDERVALUED = "G1"
PORT_CELL_SOME_UNDERVALUED = "G2"
PORT_CELL_FAIR_VALUE = "G3"
PORT_CELL_OVER_VALUED = "G4"
PORT_CELL_VALUE_NA = "G5"
PORT_COLUMN_UPSIDE = "B"
PORT_COLUMN_CUSHION = "C"
DIFF_COLUMN_UNDERVALUED = "I"
DIFF_COLUMN_SOME_UNDERVALUED = "L"
DIFF_COLUMN_FAIR = "O"
DIFF_COLUMN_OVERVALUED = "R"
DIFF_COLUMN_NA = "U"
SHEET_DIFF = "DIFF"
TIME_CELL = "T1"

def getStockValuations(ws,stocks):

    colValuation = ws[COLUMN_VALUATION]

    for cell in colValuation[1:]:
        if cell.value:
            stocks[str(cell.value)].append(cell.offset(0,-24).value) #GET SYMBOL
        else:
            break

previousResult = r'C:\Users\Yi\Dropbox\Programming\Project Watchlist\oldResult.xlsx'
oldStocksValuation = {PORT_CELL_UNDERVALUED:[],PORT_CELL_SOME_UNDERVALUED:[],PORT_CELL_FAIR_VALUE:[],PORT_CELL_OVER_VALUED:[],PORT_CELL_VALUE_NA:[]}
wsOldResult = None
currentResultFilePath = r'C:\Users\Yi\Dropbox\Programming\Project Watchlist\result.xlsx'
sourceFile = r'C:\Users\Yi\Dropbox\Programming\Project Watchlist\Watchlist.xlsx'

wb = load_workbook(filename = sourceFile,data_only = True)
ws = wb[SHEET_NAME]
wbCur = load_workbook(filename = currentResultFilePath,data_only = True)

wsPort = wb[PORT_SHEET_NAME]
colSymbol = ws[COLUMN_SYMBOL]
wb.create_sheet(RESULT_SHEET_NAME)
wsResult = wb[RESULT_SHEET_NAME]

start = wbCur[RESULT_SHEET_NAME][TIME_CELL].value
#check if we have old data
days = None
if start:
    tz = timezone('US/Eastern')
    now = datetime.datetime.now(tz)
    if now.hour > 16:
        end = datetime.date( now.year, now.month, now.day )
    else:
        end = date.today() - timedelta(1)
    while end.weekday() > 4:
        end = end - timedelta(1)#we want to get rid of weekends
    days = numpy.busday_count( start, end )

if days and days > 0:
    os.remove(previousResult)
    copyfile(currentResultFilePath, previousResult)

wbOld = load_workbook(filename = previousResult,data_only = True)

wsOldResult = wbOld[RESULT_SHEET_NAME]

getStockValuations(wsOldResult,oldStocksValuation)

for cell in colSymbol:
    if cell.value:
        symbolsCount += 1
    else:
        break

pos = []
for cell in wsPort[PORT_COLUMN_SYMBOL]:
    if cell.value:
        numOfPos += 1
        pos.append(str(cell.value))
    else:
        break
#get rid of header


del pos[0]

stocks = [Stock() for _ in range(symbolsCount)]

for i in range(0,symbolsCount):
    stocks[i].companyName = ws[COLUMN_NAME+str(i+2)].value
    print(stocks[i].companyName)
    stocks[i].symbol = ws[COLUMN_SYMBOL+str(i+2)].value
    if ws[COLUMN_BUSINESS_VALUE+str(i+2)].value : stocks[i].business = float(ws[COLUMN_BUSINESS_VALUE+str(i+2)].value)
    else :
        stocks[i].business = None
        stocks[i].fairPrice = None

    if ws[COLUMN_CASH_VALUE+str(i+2)].value:
        stocks[i].cash = float(ws[COLUMN_CASH_VALUE+str(i+2)].value)
    elif ws[COLUMN_CASH_VALUE+str(i+2)].value == 0:
        stocks[i].cash = 0.0
    else: stocks[i].cash = None

    if ws[COLUMN_CUSHION+str(i+2)].value :
        stocks[i].cushion = float(ws[COLUMN_CUSHION+str(i+2)].value)
        stocks[i].fairPrice = float(stocks[i].business+stocks[i].cash)
    else:
        stocks[i].cushion = None


    if ws[COLUMN_THESIS+str(i+2)].value : stocks[i].thesis = str(ws[COLUMN_THESIS+str(i+2)].value)
    else :stocks[i].thesis = None
    stocks[i].upside = None

wb.remove(ws)

#make the API call, chunk symbol
ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

for i in range(math.ceil(symbolsCount/BATCH_SIZE)):
    urlSymbols = ""
    for j in range(BATCH_SIZE):
        if colSymbol[i*BATCH_SIZE+j+1].value:
            urlSymbols += colSymbol[i*BATCH_SIZE+j+1].value+","
        else: break
        #loop through symbols column and concat the symbols up to the batch size limit
    urlSymbols = urlSymbols[:-1]
    url = "https://api.iextrading.com/1.0/stock/market/batch?symbols="+urlSymbols+"&types=quote&filter=symbol,close"

    data = urllib.request.urlopen(url,context=ctx).read().decode()
    js = json.loads(data)

    for k in range(BATCH_SIZE):
         if colSymbol[i*BATCH_SIZE+k+1].value:
            stocks[i*BATCH_SIZE+k].curPrice=float(js[str(stocks[i*BATCH_SIZE+k].symbol)]['quote']['close'])

sortedStocks = []
unsortedStocks = []
portStocks = {}

for i in range(0,symbolsCount):
    if stocks[i].fairPrice:
        upside = (stocks[i].fairPrice - stocks[i].curPrice)/stocks[i].curPrice
        # if upside/stocks[i].cushion >= UPSIDE_ALERT:
        if stocks[i].business/(UPSIDE_ALERT*stocks[i].cushion+1)+stocks[i].cash >= stocks[i].curPrice:
            if stocks[i].business/(stocks[i].cushion+1)+stocks[i].cash >= stocks[i].curPrice:
                if any(stock == stocks[i].symbol for stock in pos):
                    stocks[i].color = 'FF00B7E5' #Light blue
                    portStocks[stocks[i].symbol] = stocks[i]
                else:stocks[i].color = 'FF58D68D' #LIGHT GREEN
                stocks[i].valuation = PORT_CELL_UNDERVALUED #undervalued
            else:
                if any(stock == stocks[i].symbol for stock in pos):
                    stocks[i].color = 'FF2980B9' #Deep Blue
                    portStocks[stocks[i].symbol] = stocks[i]
                else:stocks[i].color = 'FF1E8449' # GREEN
                stocks[i].valuation = PORT_CELL_SOME_UNDERVALUED #somewhat undervalued
            sortedStocks.append(stocks[i])
        elif stocks[i].business/(DOWNSIDE_ALERT*stocks[i].cushion+1)+stocks[i].cash <= stocks[i].curPrice:
            if any(stock == stocks[i].symbol for stock in pos):
                stocks[i].color = 'FFE74C3C' #Crimson read
                portStocks[stocks[i].symbol] = stocks[i]
            else: stocks[i].color = 'FFF1948A' #Light RED
            stocks[i].valuation = PORT_CELL_OVER_VALUED #overvalued
            sortedStocks.append(stocks[i])
        else:
            if any(stock == stocks[i].symbol for stock in pos):
                stocks[i].color = 'FFF7DC6F' #Yellow

                portStocks[stocks[i].symbol] = stocks[i]
            else: stocks[i].color = 'FFFFFFFF' #WHITE
            stocks[i].valuation = PORT_CELL_FAIR_VALUE #fairly valued
            unsortedStocks.append(stocks[i])
        stocks[i].upside = upside
    else:
        stocks[i].color = 'FFFFFFFF' #WHITE
        unsortedStocks.append(stocks[i])
        if any(stock == stocks[i].symbol for stock in pos):
            portStocks[stocks[i].symbol] = stocks[i]
        stocks[i].valuation = PORT_CELL_VALUE_NA #valuation not available

stockValuations = {}
#populate Port Sheet header
wsPort[PORT_COLUMN_SYMBOL+str(1)].value = "SYMBOL"
wsPort[PORT_COLUMN_UPSIDE+str(1)].value = "UPSIDE"
wsPort[PORT_COLUMN_CUSHION+str(1)].value = "CUSHION"
#Color portfolio sheet
for i in range(0,numOfPos):
    wsPort[PORT_COLUMN_SYMBOL+str(i+2)].fill = PatternFill(fgColor=portStocks[pos[i]].color, fill_type = "solid")
    if portStocks[pos[i]].upside :
        wsPort[PORT_COLUMN_UPSIDE+str(i+2)].value = portStocks[pos[i]].upside
        wsPort[PORT_COLUMN_UPSIDE+str(i+2)].number_format = '0.00%'
    if portStocks[pos[i]].cushion:
        wsPort[PORT_COLUMN_CUSHION+str(i+2)].value = portStocks[pos[i]].cushion
        wsPort[PORT_COLUMN_CUSHION+str(i+2)].number_format = '0.00%'
    stockValuations[portStocks[pos[i]].valuation] = stockValuations.get(portStocks[pos[i]].valuation,0)+1

for key,value in stockValuations.items():
    wsPort[key].value = value
    wsPort[key].offset(0,1).value = value/numOfPos
    wsPort[key].offset(0,1).number_format = '0%'
sortedStocks.sort(key=lambda x: x.upside, reverse=True)


#populate header
wsResult[COLUMN_NAME+str(1)].value = "NAME"
wsResult[COLUMN_SYMBOL+str(1)].value = "SYMBOL"
wsResult[COLUMN_CUR_PRICE+str(1)].value = "CURRENT PRICE"
wsResult[COLUMN_FAIR_PRICE+str(1)].value = "FAIR PRICE"
wsResult[COLUMN_BUSINESS_VALUE+str(1)].value = "BUSINESS"
wsResult[COLUMN_CASH_VALUE+str(1)].value = "CASH"
wsResult[COLUMN_UPSIDE+str(1)].value = "UPSIDE"
wsResult[COLUMN_CUSHION+str(1)].value = "CUSHION"
wsResult[COLUMN_THESIS+str(1)].value = "THESIS"
wsResult[COLUMN_VALUATION+str(1)].value = "VALUATION"

#Populate data and Color result sheet
for i in range(len(sortedStocks)):
    wsResult[COLUMN_NAME+str(i+2)].value = sortedStocks[i].companyName
    wsResult[COLUMN_SYMBOL+str(i+2)].value = sortedStocks[i].symbol
    wsResult[COLUMN_CUR_PRICE+str(i+2)].value = sortedStocks[i].curPrice
    if sortedStocks[i].fairPrice:
        wsResult[COLUMN_FAIR_PRICE+str(i+2)].value = sortedStocks[i].fairPrice
        wsResult[COLUMN_FAIR_PRICE+str(i+2)].number_format = '0.00'
        wsResult[COLUMN_UPSIDE+str(i+2)].value = sortedStocks[i].upside
        wsResult[COLUMN_UPSIDE+str(i+2)].number_format = '0.00%'
        wsResult[COLUMN_BUSINESS_VALUE+str(i+2)].value = sortedStocks[i].business
        wsResult[COLUMN_BUSINESS_VALUE+str(i+2)].number_format = '0.00'
        wsResult[COLUMN_CASH_VALUE+str(i+2)].value = sortedStocks[i].cash
        wsResult[COLUMN_CASH_VALUE+str(i+2)].number_format = '0.00'
        wsResult[COLUMN_CUSHION+str(i+2)].value = sortedStocks[i].cushion
        wsResult[COLUMN_CUSHION+str(i+2)].number_format = '0.00%'
    if sortedStocks[i].thesis: wsResult[COLUMN_THESIS+str(i+2)].value = sortedStocks[i].thesis
    for cell in wsResult[str(i+2)+":"+str(i+2)]:
        cell.fill = PatternFill(fgColor=sortedStocks[i].color, fill_type = "solid")
    wsResult[COLUMN_VALUATION+str(i+2)].value = sortedStocks[i].valuation

for i in range(len(sortedStocks),len(unsortedStocks)+len(sortedStocks)):
    wsResult[COLUMN_NAME+str(i+2)].value = unsortedStocks[i-len(sortedStocks)].companyName
    wsResult[COLUMN_SYMBOL+str(i+2)].value = unsortedStocks[i-len(sortedStocks)].symbol
    wsResult[COLUMN_CUR_PRICE+str(i+2)].value = unsortedStocks[i-len(sortedStocks)].curPrice
    if unsortedStocks[i-len(sortedStocks)].thesis: wsResult[COLUMN_THESIS+str(i+2)].value = unsortedStocks[i-len(sortedStocks)].thesis
    if unsortedStocks[i-len(sortedStocks)].fairPrice:
        wsResult[COLUMN_FAIR_PRICE+str(i+2)].value = unsortedStocks[i-len(sortedStocks)].fairPrice
        wsResult[COLUMN_FAIR_PRICE+str(i+2)].number_format = '0.00'
        wsResult[COLUMN_CUSHION+str(i+2)].value = unsortedStocks[i-len(sortedStocks)].cushion
        wsResult[COLUMN_CUSHION+str(i+2)].number_format = '0.00%'
        wsResult[COLUMN_UPSIDE+str(i+2)].value = unsortedStocks[i-len(sortedStocks)].upside
        wsResult[COLUMN_UPSIDE+str(i+2)].number_format = '0.00%'
        wsResult[COLUMN_BUSINESS_VALUE+str(i+2)].value = unsortedStocks[i-len(sortedStocks)].business
        wsResult[COLUMN_BUSINESS_VALUE+str(i+2)].number_format = '0.00'
        wsResult[COLUMN_CASH_VALUE+str(i+2)].value = unsortedStocks[i-len(sortedStocks)].cash
        wsResult[COLUMN_CASH_VALUE+str(i+2)].number_format = '0.00'
    for cell in wsResult[str(i+2)+":"+str(i+2)]:
        cell.fill = PatternFill(fgColor=unsortedStocks[i-len(sortedStocks)].color, fill_type = "solid")
    wsResult[COLUMN_VALUATION+str(i+2)].value = unsortedStocks[i-len(sortedStocks)].valuation

curStocksValuation = {PORT_CELL_UNDERVALUED:[],PORT_CELL_SOME_UNDERVALUED:[],PORT_CELL_FAIR_VALUE:[],PORT_CELL_OVER_VALUED:[],PORT_CELL_VALUE_NA:[]}

getStockValuations(wsResult,curStocksValuation)

wsResult[TIME_CELL].value = str(end)
#POPULATE DIFF SHEET
wb.create_sheet(SHEET_DIFF)
wsDiff= wb[SHEET_DIFF]
#POPULATE HEADERS
wsDiff[DIFF_COLUMN_UNDERVALUED+str(1)].value = "UNDERVALUED"
wsDiff[DIFF_COLUMN_UNDERVALUED+str(1)].offset(0,1).value = "IN"
wsDiff[DIFF_COLUMN_UNDERVALUED+str(1)].offset(0,2).value = "OUT"
for cell in wsDiff[DIFF_COLUMN_UNDERVALUED+":"+DIFF_COLUMN_UNDERVALUED]:
    cell.fill = PatternFill(fgColor='FF58D68D', fill_type = "solid")

wsDiff[DIFF_COLUMN_SOME_UNDERVALUED+str(1)].value = "SLIGHTLY UNDERVALUED"
wsDiff[DIFF_COLUMN_SOME_UNDERVALUED+str(1)].offset(0,1).value = "IN"
wsDiff[DIFF_COLUMN_SOME_UNDERVALUED+str(1)].offset(0,2).value = "OUT"
for cell in wsDiff[DIFF_COLUMN_SOME_UNDERVALUED+":"+DIFF_COLUMN_SOME_UNDERVALUED]:
    cell.fill = PatternFill(fgColor='FF1E8449', fill_type = "solid")

wsDiff[DIFF_COLUMN_FAIR+str(1)].value = "FAIR"
wsDiff[DIFF_COLUMN_FAIR+str(1)].offset(0,1).value = "IN"
wsDiff[DIFF_COLUMN_FAIR+str(1)].offset(0,2).value = "OUT"
for cell in wsDiff[DIFF_COLUMN_FAIR+":"+DIFF_COLUMN_FAIR]:
    cell.fill = PatternFill(fgColor='FFF1948A', fill_type = "solid")

wsDiff[DIFF_COLUMN_OVERVALUED+str(1)].value = "OVERVALUED"
wsDiff[DIFF_COLUMN_OVERVALUED+str(1)].offset(0,1).value = "IN"
wsDiff[DIFF_COLUMN_OVERVALUED+str(1)].offset(0,2).value = "OUT"
for cell in wsDiff[DIFF_COLUMN_OVERVALUED+":"+DIFF_COLUMN_OVERVALUED]:
    cell.fill = PatternFill(fgColor='FFF7DC6F', fill_type = "solid")

wsDiff[DIFF_COLUMN_NA+str(1)].value = "N/A"
wsDiff[DIFF_COLUMN_NA+str(1)].offset(0,1).value = "IN"
wsDiff[DIFF_COLUMN_NA+str(1)].offset(0,2).value = "OUT"
for cell in wsDiff[DIFF_COLUMN_NA+":"+DIFF_COLUMN_NA]:
    cell.fill = PatternFill(fgColor='FF58D68D', fill_type = "solid")



def populateDiffColumnData(stocks,anchorCell):
    if stocks :
        for i in range(len(stocks)):
            anchorCell.offset(i+1).value = stocks[i]

outUnderValuedStocks = list(set(oldStocksValuation[PORT_CELL_UNDERVALUED])-set(curStocksValuation[PORT_CELL_UNDERVALUED]))
populateDiffColumnData(outUnderValuedStocks,wsDiff[DIFF_COLUMN_UNDERVALUED+str(1)].offset(0,2))
inUnderValuedStocks = list(set(curStocksValuation[PORT_CELL_UNDERVALUED])-set(oldStocksValuation[PORT_CELL_UNDERVALUED]))
populateDiffColumnData(inUnderValuedStocks,wsDiff[DIFF_COLUMN_UNDERVALUED+str(1)].offset(0,1))

outSUnderValuedStocks = list(set(oldStocksValuation[PORT_CELL_SOME_UNDERVALUED])-set(curStocksValuation[PORT_CELL_SOME_UNDERVALUED]))
populateDiffColumnData(outSUnderValuedStocks,wsDiff[DIFF_COLUMN_SOME_UNDERVALUED+str(1)].offset(0,2))
inSUnderValuedStocks = list(set(curStocksValuation[PORT_CELL_SOME_UNDERVALUED])-set(oldStocksValuation[PORT_CELL_SOME_UNDERVALUED]))
populateDiffColumnData(inSUnderValuedStocks,wsDiff[DIFF_COLUMN_SOME_UNDERVALUED+str(1)].offset(0,1))

outFairValuedStocks = list(set(oldStocksValuation[PORT_CELL_FAIR_VALUE])-set(curStocksValuation[PORT_CELL_FAIR_VALUE]))
populateDiffColumnData(outFairValuedStocks,wsDiff[DIFF_COLUMN_FAIR+str(1)].offset(0,2))
inFairValuedStocks = list(set(curStocksValuation[PORT_CELL_FAIR_VALUE])-set(oldStocksValuation[PORT_CELL_FAIR_VALUE]))
populateDiffColumnData(inFairValuedStocks,wsDiff[DIFF_COLUMN_FAIR+str(1)].offset(0,1))

outOvervaluedStocks = list(set(oldStocksValuation[PORT_CELL_OVER_VALUED])-set(curStocksValuation[PORT_CELL_OVER_VALUED]))
populateDiffColumnData(outOvervaluedStocks,wsDiff[DIFF_COLUMN_OVERVALUED+str(1)].offset(0,2))
inOvervaluedStocks = list(set(curStocksValuation[PORT_CELL_OVER_VALUED])-set(oldStocksValuation[PORT_CELL_OVER_VALUED]))
populateDiffColumnData(inOvervaluedStocks,wsDiff[DIFF_COLUMN_OVERVALUED+str(1)].offset(0,1))

outNAStocks = list(set(oldStocksValuation[PORT_CELL_VALUE_NA])-set(curStocksValuation[PORT_CELL_VALUE_NA]))
populateDiffColumnData(outNAStocks,wsDiff[DIFF_COLUMN_NA+str(1)].offset(0,2))
inNAStocks = list(set(curStocksValuation[PORT_CELL_VALUE_NA])-set(oldStocksValuation[PORT_CELL_VALUE_NA]))
populateDiffColumnData(inNAStocks,wsDiff[DIFF_COLUMN_NA+str(1)].offset(0,1))

#compare cur valuation and old valuation
#TODO NICE TO HAVE PAINT WHOLE CATEGORY COLUMNS

wb.save(filename = currentResultFilePath)

#SEND EMAIL

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

if  inUnderValuedStocks or inSUnderValuedStocks or  inFairValuedStocks or  inOvervaluedStocks or   inNAStocks:
    email_user = 'globalvaluescanner@gmail.com'
    email_password = 'PythonEmailService!163'
    email_send = '17714346535@163.com'

    subject = 'Valuation Update'

    msg = MIMEMultipart()
    msg['From'] = email_user
    msg['To'] = email_send
    msg['Subject'] = subject

    body = 'Position Update'
    msg.attach(MIMEText(body,'plain'))

    filename=currentResultFilePath
    attachment  =open(filename,'rb')

    part = MIMEBase('application','octet-stream')
    part.set_payload((attachment).read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition',"attachment; filename= "+os.path.basename(filename))

    msg.attach(part)
    text = msg.as_string()
    server = smtplib.SMTP('smtp.gmail.com',587)
    server.starttls()
    server.login(email_user,email_password)


    server.sendmail(email_user,email_send,text)
    server.quit()
