from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import math
import urllib.request,urllib.parse, urllib.error
import ssl
from ClassStock import Stock
import json
import os, errno

#TODO TODO separate Business valuation and Cash equivalent valuation so the calculation is correct
#TODO nice to have: put focus back on the result sheet instead of port sheet
#TODO nice to have: compare current snapshot with historical data, might be able to try this with tableau
#TODO nice to have: pie chart for port allocation

previousResult = r'C:\Users\Yi\Dropbox\Programming\Project Watchlist\result.xlsx'
try:
    os.remove(previousResult)
except OSError:
    pass

wb = load_workbook(filename = r'C:\Users\Yi\Dropbox\Programming\Project Watchlist\Watchlist.xlsx',data_only = True)

COLUMN_NAME = 'A'
COLUMN_SYMBOL = 'B'
COLUMN_CUR_PRICE = 'C'
COLUMN_FAIR_PRICE = 'D'
COLUMN_THESIS = 'E'
COLUMN_UPSIDE = 'F'
COLUMN_CUSHION = 'G'
SHEET_NAME = "Sheet1"
PORT_SHEET_NAME = "Port"
RESULT_SHEET_NAME = "Result"
PORT_COLUMN_SYMBOL = 'A'
ws = wb[SHEET_NAME]
wsPort = wb[PORT_SHEET_NAME]
colSymbol = ws[COLUMN_SYMBOL]
BATCH_SIZE = 100
UPSIDE_ALERT = 0.9
DOWNSIDE_ALERT = 0.15
#count number of rows
symbolsCount = -1#offset the header count
numOfPos  = -1
PORT_CELL_UNDERVALUED = "G1"
PORT_CELL_SOME_UNDERVALUED = "G2"
PORT_CELL_FAIR_VALUE = "G3"
PORT_CELL_OVER_VALUED = "G4"
PORT_CELL_VALUE_NA = "G5"

wb.create_sheet(RESULT_SHEET_NAME)
wsResult = wb[RESULT_SHEET_NAME]

for cell in colSymbol:
    if cell.value:
        symbolsCount += 1
    else:
        break

pos = []
for cell in wsPort[PORT_COLUMN_SYMBOL]:
    if cell.value:
        numOfPos += 1
        pos.append(str(cell.value))
    else:
        break
#get rid of header
del pos[0]

stocks = [Stock() for _ in range(symbolsCount)]

for i in range(0,symbolsCount):
    stocks[i].companyName = ws[COLUMN_NAME+str(i+2)].value
    stocks[i].symbol = ws[COLUMN_SYMBOL+str(i+2)].value
    if ws[COLUMN_FAIR_PRICE+str(i+2)].value : stocks[i].fairPrice = float(ws[COLUMN_FAIR_PRICE+str(i+2)].value)
    else :stocks[i].fairPrice = None
    if ws[COLUMN_CUSHION+str(i+2)].value : stocks[i].cushion = float(ws[COLUMN_CUSHION+str(i+2)].value)
    if ws[COLUMN_THESIS+str(i+2)].value : stocks[i].thesis = str(ws[COLUMN_THESIS+str(i+2)].value)
    else :stocks[i].thesis = None

wb.remove(ws)

#make the API call, chunk symbol
ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

for i in range(math.ceil(symbolsCount/BATCH_SIZE)):
    urlSymbols = ""
    for j in range(BATCH_SIZE):
        if colSymbol[i*BATCH_SIZE+j+1].value:
            urlSymbols += colSymbol[i*BATCH_SIZE+j+1].value+","
        else: break
        #loop through symbols column and concat the symbols up to the batch size limit
    urlSymbols = urlSymbols[:-1]
    url = "https://api.iextrading.com/1.0/stock/market/batch?symbols="+urlSymbols+"&types=quote&filter=symbol,close"

    data = urllib.request.urlopen(url,context=ctx).read().decode()
    js = json.loads(data)

    for k in range(BATCH_SIZE):
         if colSymbol[i*BATCH_SIZE+k+1].value:
            stocks[i*BATCH_SIZE+k].curPrice=float(js[str(stocks[i*BATCH_SIZE+k].symbol)]['quote']['close'])

sortedStocks = []
unsortedStocks = []
portStocks = {}
for i in range(0,symbolsCount):
    if stocks[i].fairPrice:
        upside = (stocks[i].fairPrice - stocks[i].curPrice)/stocks[i].curPrice
        if upside/stocks[i].cushion >= UPSIDE_ALERT:
            if upside > stocks[i].cushion:
                if any(stock == stocks[i].symbol for stock in pos):
                    stocks[i].color = 'FF00B7E5' #Light blue
                    stocks[i].valuation = PORT_CELL_UNDERVALUED #undervalued
                    portStocks[stocks[i].symbol] = stocks[i]
                else:stocks[i].color = 'FF58D68D' #LIGHT GREEN
            else:
                if any(stock == stocks[i].symbol for stock in pos):
                    stocks[i].color = 'FF2980B9' #Deep Blue
                    stocks[i].valuation = PORT_CELL_SOME_UNDERVALUED #somewhat undervalued
                    portStocks[stocks[i].symbol] = stocks[i]
                else:stocks[i].color = 'FF1E8449' # GREEN
            sortedStocks.append(stocks[i])
        elif upside <= DOWNSIDE_ALERT:
            if any(stock == stocks[i].symbol for stock in pos):
                stocks[i].color = 'FFE74C3C' #Crimson read
                stocks[i].valuation = PORT_CELL_OVER_VALUED #overvalued
                portStocks[stocks[i].symbol] = stocks[i]
            else: stocks[i].color = 'FFF1948A' #Light RED
            sortedStocks.append(stocks[i])
        else:
            if any(stock == stocks[i].symbol for stock in pos):
                stocks[i].color = 'FFF7DC6F' #Yellow
                stocks[i].valuation = PORT_CELL_FAIR_VALUE #fairly valued
                portStocks[stocks[i].symbol] = stocks[i]
            else: stocks[i].color = 'FFFFFFFF' #WHITE
            unsortedStocks.append(stocks[i])
        stocks[i].upside = upside
    else:
        stocks[i].color = 'FFFFFFFF' #WHITE
        unsortedStocks.append(stocks[i])
        if any(stock == stocks[i].symbol for stock in pos):
            stocks[i].valuation = PORT_CELL_VALUE_NA #valuation not available
            portStocks[stocks[i].symbol] = stocks[i]

stockValuations = {}
#Color portfolio sheet
for i in range(0,numOfPos):
    wsPort[PORT_COLUMN_SYMBOL+str(i+2)].fill = PatternFill(fgColor=portStocks[pos[i]].color, fill_type = "solid")
    stockValuations[portStocks[pos[i]].valuation] = stockValuations.get(portStocks[pos[i]].valuation,0)+1

for key,value in stockValuations.items():
    wsPort[key].value = value
    wsPort[key].offset(0,1).value = value/numOfPos
    wsPort[key].offset(0,1).number_format = '0%'
sortedStocks.sort(key=lambda x: x.upside, reverse=True)


#populate header
wsResult[COLUMN_NAME+str(1)].value = "NAME"
wsResult[COLUMN_SYMBOL+str(1)].value = "SYMBOL"
wsResult[COLUMN_CUR_PRICE+str(1)].value = "CURRENT PRICE"
wsResult[COLUMN_FAIR_PRICE+str(1)].value = "FAIR PRICE"
wsResult[COLUMN_UPSIDE+str(1)].value = "UPSIDE"
wsResult[COLUMN_CUSHION+str(1)].value = "CUSHION"
wsResult[COLUMN_THESIS+str(1)].value = "THESIS"

#Populate data and Color result sheet
for i in range(len(sortedStocks)):
    wsResult[COLUMN_NAME+str(i+2)].value = sortedStocks[i].companyName
    wsResult[COLUMN_SYMBOL+str(i+2)].value = sortedStocks[i].symbol
    wsResult[COLUMN_CUR_PRICE+str(i+2)].value = sortedStocks[i].curPrice
    if sortedStocks[i].fairPrice:
        wsResult[COLUMN_FAIR_PRICE+str(i+2)].value = sortedStocks[i].fairPrice
        wsResult[COLUMN_FAIR_PRICE+str(i+2)].number_format = '0.00'
        wsResult[COLUMN_UPSIDE+str(i+2)].value = sortedStocks[i].upside
        wsResult[COLUMN_UPSIDE+str(i+2)].number_format = '0.00%'
        wsResult[COLUMN_CUSHION+str(i+2)].value = sortedStocks[i].cushion
        wsResult[COLUMN_CUSHION+str(i+2)].number_format = '0.00%'
    if sortedStocks[i].thesis: wsResult[COLUMN_THESIS+str(i+2)].value = sortedStocks[i].thesis
    for cell in wsResult[str(i+2)+":"+str(i+2)]:
        cell.fill = PatternFill(fgColor=sortedStocks[i].color, fill_type = "solid")

for i in range(len(sortedStocks),len(unsortedStocks)+len(sortedStocks)):
    wsResult[COLUMN_NAME+str(i+2)].value = unsortedStocks[i-len(sortedStocks)].companyName
    wsResult[COLUMN_SYMBOL+str(i+2)].value = unsortedStocks[i-len(sortedStocks)].symbol
    wsResult[COLUMN_CUR_PRICE+str(i+2)].value = unsortedStocks[i-len(sortedStocks)].curPrice
    if unsortedStocks[i-len(sortedStocks)].thesis: wsResult[COLUMN_THESIS+str(i+2)].value = unsortedStocks[i-len(sortedStocks)].thesis
    if unsortedStocks[i-len(sortedStocks)].fairPrice:
        wsResult[COLUMN_FAIR_PRICE+str(i+2)].value = unsortedStocks[i-len(sortedStocks)].fairPrice
        wsResult[COLUMN_FAIR_PRICE+str(i+2)].number_format = '0.00'
        wsResult[COLUMN_CUSHION+str(i+2)].value = unsortedStocks[i-len(sortedStocks)].cushion
        wsResult[COLUMN_CUSHION+str(i+2)].number_format = '0.00%'
        wsResult[COLUMN_UPSIDE+str(i+2)].value = unsortedStocks[i-len(sortedStocks)].upside
        wsResult[COLUMN_UPSIDE+str(i+2)].number_format = '0.00%'
    for cell in wsResult[str(i+2)+":"+str(i+2)]:
        cell.fill = PatternFill(fgColor=unsortedStocks[i-len(sortedStocks)].color, fill_type = "solid")

wb.save(filename = r'C:\Users\Yi\Dropbox\Programming\Project Watchlist\result.xlsx')
